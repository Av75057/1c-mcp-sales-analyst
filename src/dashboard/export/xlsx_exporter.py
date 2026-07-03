from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def export_xlsx(data: list[dict[str, Any]], meta: dict[str, Any] | None = None) -> io.BytesIO:
    wb = Workbook()

    # Sheet 1: Metadata
    ws_meta = wb.active
    ws_meta.title = "Info"
    ws_meta.append(["Title", meta.get("title", "") if meta else ""])
    ws_meta.append(["Description", meta.get("description", "") if meta else ""])
    ws_meta.append(["Query", meta.get("original_query", "") if meta else ""])
    ws_meta.append(["Generated", datetime.utcnow().isoformat()])
    for row in ws_meta.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True)

    # Sheet 2: Data
    if data:
        ws_data = wb.create_sheet("Data")
        cols = list(data[0].keys())
        ws_data.append(cols)
        hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws_data[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")
        for row in data:
            ws_data.append([row.get(c, "") for c in cols])
        for col in ws_data.columns:
            ml = max(len(str(c.value or "")) for c in col)
            ws_data.column_dimensions[col[0].column_letter].width = min(ml + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
